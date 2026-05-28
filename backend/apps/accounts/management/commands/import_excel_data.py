from collections import defaultdict
from datetime import datetime, time, timedelta
from decimal import Decimal, InvalidOperation
import warnings

from django.contrib.auth import get_user_model
from django.core.management.base import BaseCommand, CommandError
from django.db import transaction
from django.utils import timezone
from openpyxl import load_workbook

from apps.accounts.management.commands.seed_demo import DEMO_USERNAMES, ROLE_LABELS
from apps.accounts.models import Role, UserProfile
from apps.clients.models import Client
from apps.common.constants import (
    ALL_ROLES,
    ROLE_ACCOUNTANT,
    ROLE_CASHIER,
    ROLE_DRIVER,
    ROLE_OWNER,
    ROLE_TECHNICIAN,
)
from apps.drivers.models import Driver, DriverPayment, DriverTrip
from apps.finance.models import (
    CashTransaction,
    Cashbox,
    Expense,
    ExpenseCategory,
    Invoice,
    InvoiceItem,
    Payment,
    PaymentMethod,
)
from apps.inventory.models import Item, ItemCategory
from apps.orders.models import JobDriverAssignment, JobStaffAssignment, Order, OrderItem
from apps.payroll.models import PayrollLine, PayrollPeriod
from apps.staff.models import StaffAdvance, StaffProfile, StaffTask
from apps.vendors.models import Vendor, VendorPayment, VendorTransaction


def clean_text(value):
    if value is None:
        return ""
    return str(value).strip()


def money(value):
    if value in (None, ""):
        return Decimal("0")
    try:
        return Decimal(str(value)).quantize(Decimal("0.01"))
    except (InvalidOperation, ValueError):
        return Decimal("0")


def as_int(value, default=0):
    try:
        return int(Decimal(str(value)))
    except (InvalidOperation, ValueError, TypeError):
        return default


def as_date(value):
    if isinstance(value, datetime):
        return value.date()
    return value


def as_aware(value):
    if isinstance(value, datetime):
        if timezone.is_naive(value):
            return timezone.make_aware(value)
        return value
    if value:
        return timezone.make_aware(datetime.combine(value, time(hour=9)))
    return timezone.now()


def row_dict(headers, row):
    return {clean_text(headers[index]): row[index] for index in range(min(len(headers), len(row)))}


class Command(BaseCommand):
    help = "Import the Arabic Excel control workbook into EventOps models."

    def add_arguments(self, parser):
        parser.add_argument("workbook", help="Path to the Excel workbook")
        parser.add_argument(
            "--clear-excel",
            action="store_true",
            help="Clear records with Excel/import codes before importing again.",
        )

    def handle(self, *args, **options):
        workbook_path = options["workbook"]
        warnings.filterwarnings("ignore", category=UserWarning, module="openpyxl")
        try:
            workbook = load_workbook(workbook_path, data_only=True)
        except FileNotFoundError as exc:
            raise CommandError(f"Workbook not found: {workbook_path}") from exc

        with transaction.atomic():
            users = self.ensure_users()
            if options["clear_excel"]:
                self.clear_excel_records()
            inventory = self.import_inventory(workbook)
            staff = self.import_staff_and_payroll(workbook, users)
            drivers = self.import_drivers(workbook, users)
            vendors = self.import_vendors(workbook)
            orders = self.import_orders(workbook, users, inventory)
            self.import_technician_costs(workbook, users, orders, staff)
            self.import_collections(workbook, users, orders)
            self.import_vendor_driver_payments(workbook, users, vendors, drivers)
            self.import_cash_custody(workbook, users, orders)

        self.stdout.write(
            self.style.SUCCESS(
                "Excel import complete: "
                f"{Client.objects.count()} clients, {Order.objects.count()} orders, "
                f"{Item.objects.count()} items, {Payment.objects.count()} payments."
            )
        )

    def ensure_users(self):
        User = get_user_model()
        users = {}
        for code in ALL_ROLES:
            name_en, name_ar = ROLE_LABELS[code]
            role, _ = Role.objects.update_or_create(
                code=code,
                defaults={
                    "name_en": name_en,
                    "name_ar": name_ar,
                    "can_view_profit": code in [ROLE_OWNER, ROLE_ACCOUNTANT],
                    "can_view_payroll": code in [ROLE_OWNER, ROLE_ACCOUNTANT],
                    "can_approve": code == ROLE_OWNER,
                },
            )
            username = DEMO_USERNAMES[code]
            user, created = User.objects.get_or_create(
                username=username,
                defaults={
                    "email": f"{username}@eventops.local",
                    "role": role,
                    "is_staff": code == ROLE_OWNER,
                    "is_superuser": code == ROLE_OWNER,
                },
            )
            if created:
                user.set_password("eventops123")
                user.save()
            elif user.role_id != role.id:
                user.role = role
                user.save(update_fields=["role"])
            UserProfile.objects.get_or_create(user=user)
            users[code] = user
        return users

    def clear_excel_records(self):
        Payment.objects.filter(reference__startswith="EXCEL-").delete()
        CashTransaction.objects.filter(description__startswith="[Excel]").delete()
        Expense.objects.filter(description__startswith="[Excel]").delete()
        Invoice.objects.filter(code__startswith="INV-EXCEL-").delete()
        StaffTask.objects.filter(notes__startswith="Excel import").delete()
        DriverTrip.objects.filter(notes__startswith="Excel import").delete()
        Order.objects.filter(code__startswith="EXCEL-").delete()

    def import_inventory(self, workbook):
        items = {}
        if "متاح" not in workbook.sheetnames:
            return items
        category, _ = ItemCategory.objects.get_or_create(name="Excel Screen Inventory")
        for row in workbook["متاح"].iter_rows(min_row=2, values_only=True):
            code = clean_text(row[0])
            if not code:
                continue
            total = as_int(row[1])
            pulled = as_int(row[2])
            available = as_int(row[3], max(total - pulled, 0))
            item, _ = Item.objects.update_or_create(
                sku=code,
                defaults={
                    "name": code,
                    "category": category,
                    "unit": "m2" if not code.startswith("خارجي") else "service",
                    "total_quantity": max(total, available),
                    "reserved_quantity": max(pulled, 0),
                    "available_quantity": max(available, 0),
                    "rental_price": Decimal("0"),
                    "location": "Excel warehouse",
                },
            )
            items[code] = item
        return items

    def import_staff_and_payroll(self, workbook, users):
        staff_by_name = {}
        if "مرتبات" in workbook.sheetnames:
            ws = workbook["مرتبات"]
            period_month = as_int(ws["E1"].value, timezone.localdate().month)
            start_date = datetime(2026, period_month, 1).date()
            end_date = (datetime(2026, period_month + 1, 1).date() - timedelta(days=1)) if period_month < 12 else datetime(2026, 12, 31).date()
            period, _ = PayrollPeriod.objects.get_or_create(
                start_date=start_date,
                end_date=end_date,
                defaults={"name": f"Excel Payroll {period_month}/2026", "status": "calculated"},
            )
            for row in ws.iter_rows(min_row=3, values_only=True):
                name = clean_text(row[0])
                if not name:
                    continue
                base_salary = money(row[1])
                variable = money(row[2])
                advances = money(row[3])
                net_pay = money(row[4])
                staff, _ = StaffProfile.objects.update_or_create(
                    name=name,
                    defaults={
                        "staff_role": "Excel technician",
                        "base_salary": base_salary,
                        "day_rate": Decimal("0"),
                    },
                )
                staff_by_name[name] = staff
                PayrollLine.objects.update_or_create(
                    period=period,
                    staff=staff,
                    defaults={
                        "base_salary": base_salary,
                        "task_pay": variable,
                        "advances": advances,
                        "net_pay": net_pay,
                        "status": "calculated",
                    },
                )
                if advances:
                    StaffAdvance.objects.get_or_create(
                        staff=staff,
                        amount=advances,
                        advance_date=start_date,
                        defaults={"notes": "Excel payroll withdrawals", "created_by": users[ROLE_ACCOUNTANT]},
                    )
        if "⚙️" in workbook.sheetnames:
            for row in workbook["⚙️"].iter_rows(min_row=2, values_only=True):
                name = clean_text(row[0])
                if name and name not in staff_by_name:
                    staff, _ = StaffProfile.objects.get_or_create(
                        name=name,
                        defaults={"staff_role": "Excel technician"},
                    )
                    staff_by_name[name] = staff
        return staff_by_name

    def import_drivers(self, workbook, users):
        drivers = {}
        if "⚙️" in workbook.sheetnames:
            for row in workbook["⚙️"].iter_rows(min_row=2, values_only=True):
                name = clean_text(row[2])
                if name:
                    driver, _ = Driver.objects.get_or_create(name=name, defaults={"day_rate": Decimal("0")})
                    drivers[name] = driver
        if "IN" in workbook.sheetnames:
            for row in workbook["IN"].iter_rows(min_row=2, values_only=True):
                for index in (22, 24):
                    name = clean_text(row[index])
                    if name:
                        driver, _ = Driver.objects.get_or_create(name=name, defaults={"day_rate": Decimal("0")})
                        drivers[name] = driver
        return drivers

    def import_vendors(self, workbook):
        vendors = {}
        if "⚙️" in workbook.sheetnames:
            for row in workbook["⚙️"].iter_rows(min_row=2, values_only=True):
                name = clean_text(row[9])
                if name:
                    vendor, _ = Vendor.objects.get_or_create(name=name, defaults={"service_type": "Excel vendor"})
                    vendors[name] = vendor
        if "الموردين" in workbook.sheetnames:
            for row in workbook["الموردين"].iter_rows(min_row=3, values_only=True):
                for index in (0, 7, 15):
                    name = clean_text(row[index])
                    if name:
                        vendor, _ = Vendor.objects.get_or_create(name=name, defaults={"service_type": "Excel vendor"})
                        vendors[name] = vendor
        return vendors

    def import_orders(self, workbook, users, inventory):
        orders = {}
        if "IN" not in workbook.sheetnames:
            return orders

        grouped = defaultdict(list)
        headers = [cell.value for cell in workbook["IN"][1]]
        for row in workbook["IN"].iter_rows(min_row=2, values_only=True):
            data = row_dict(headers, row)
            order_id = clean_text(data.get("أمر \nID") or data.get("أمر ID"))
            client_name = clean_text(data.get("العميل"))
            if not order_id or not client_name or "Total" in client_name:
                continue
            grouped[order_id].append(data)

        imported_category, _ = ItemCategory.objects.get_or_create(name="Excel Imported Codes")
        for order_id, lines in grouped.items():
            first = lines[0]
            client, _ = Client.objects.get_or_create(name=clean_text(first.get("العميل")))
            start_at = as_aware(first.get("التاريخ"))
            end_at = as_aware(first.get("تاريخ \nالانتهاء") or first.get("تاريخ الانتهاء")) or (start_at + timedelta(days=1))
            if end_at <= start_at:
                end_at = start_at + timedelta(days=1)
            total_amount = sum((money(line.get("المستحق 🔼")) for line in lines), Decimal("0"))
            collected = sum((money(line.get("المُحصل 🔼")) for line in lines), Decimal("0"))
            revenue = sum((money(line.get("الصافي 🔼")) for line in lines), Decimal("0"))
            remaining = max(total_amount - collected, Decimal("0"))
            status = "fully_paid" if total_amount and remaining == 0 else "partially_paid" if collected else "completed"
            order, _ = Order.objects.update_or_create(
                code=f"EXCEL-{order_id}",
                defaults={
                    "client": client,
                    "title": f"{client.name} - {clean_text(first.get('الموقع')) or order_id}",
                    "event_location": clean_text(first.get("الموقع")),
                    "start_at": start_at,
                    "end_at": end_at,
                    "status": status,
                    "payment_status": "fully_paid" if total_amount and remaining == 0 else "partially_paid" if collected else "unpaid",
                    "revenue_amount": revenue,
                    "discount_amount": Decimal("0"),
                    "tax_amount": Decimal("0"),
                    "total_amount": total_amount,
                    "notes": "Imported from Excel sheet IN",
                    "created_by": users[ROLE_ACCOUNTANT],
                },
            )
            orders[order_id] = order
            OrderItem.objects.filter(order=order).delete()
            for line in lines:
                sku = clean_text(line.get("الكود")) or "UNKNOWN"
                item = inventory.get(sku)
                if item is None:
                    item, _ = Item.objects.get_or_create(
                        sku=sku,
                        defaults={
                            "name": sku,
                            "category": imported_category,
                            "unit": "m2",
                            "total_quantity": as_int(line.get("إجمالى أمتار"), 0),
                            "available_quantity": as_int(line.get("المتاح"), 0),
                        },
                    )
                    inventory[sku] = item
                line_total = money(line.get("المستحق 🔼"))
                net_profit = money(line.get("صافى الاوردر"))
                cost_total = max(line_total - net_profit, Decimal("0"))
                quantity = as_int(line.get("إجمالى أمتار"), 1) or as_int(line.get("M2"), 1) or 1
                OrderItem.objects.create(
                    order=order,
                    item=item,
                    description=f"Excel code {sku}",
                    quantity=max(quantity, 1),
                    unit_price=money(line.get("سعر\nالمتر 🔼")),
                    cost_price=cost_total,
                    start_at=start_at,
                    end_at=end_at,
                )
                for driver_index, cost_index, label in ((22, 23, "delivery"), (24, 25, "return")):
                    driver_name = clean_text(list(line.values())[driver_index]) if len(line.values()) > driver_index else ""
                    driver_cost = money(list(line.values())[cost_index]) if len(line.values()) > cost_index else Decimal("0")
                    if driver_name and driver_cost:
                        driver, _ = Driver.objects.get_or_create(name=driver_name, defaults={"day_rate": Decimal("0")})
                        JobDriverAssignment.objects.get_or_create(
                            order=order,
                            driver=driver,
                            scheduled_at=start_at,
                            pickup_location="Excel",
                            dropoff_location=clean_text(first.get("الموقع")),
                            defaults={"cost": driver_cost, "status": label},
                        )
                        DriverTrip.objects.get_or_create(
                            driver=driver,
                            order=order,
                            pickup_location="Excel",
                            dropoff_location=clean_text(first.get("الموقع")),
                            scheduled_at=start_at,
                            defaults={"status": "completed", "cost": driver_cost, "notes": "Excel import driver trip"},
                        )
                vendor_name = clean_text(line.get("مورد خارجي 🔽"))
                vendor_cost = money(line.get("اضافى 🔽"))
                if vendor_name:
                    vendor, _ = Vendor.objects.get_or_create(name=vendor_name, defaults={"service_type": "Excel external vendor"})
                    VendorTransaction.objects.get_or_create(
                        vendor=vendor,
                        order=order,
                        description=f"Excel external vendor for {order.code}",
                        defaults={
                            "transaction_type": "cost",
                            "amount": vendor_cost,
                            "transaction_date": as_date(first.get("التاريخ")),
                            "created_by": users[ROLE_ACCOUNTANT],
                        },
                    )

            invoice, _ = Invoice.objects.update_or_create(
                code=f"INV-EXCEL-{order_id}",
                defaults={
                    "client": client,
                    "order": order,
                    "issue_date": as_date(first.get("التاريخ")),
                    "due_date": as_date(first.get("تاريخ \nالاستحقاق") or first.get("تاريخ الاستحقاق")),
                    "status": "paid" if remaining == 0 and total_amount else "partially_paid" if collected else "issued",
                    "subtotal": total_amount,
                    "total_amount": total_amount,
                    "created_by": users[ROLE_ACCOUNTANT],
                    "notes": "Imported from Excel workbook",
                },
            )
            InvoiceItem.objects.filter(invoice=invoice).delete()
            InvoiceItem.objects.create(
                invoice=invoice,
                description=f"Excel order {order_id}",
                quantity=Decimal("1"),
                unit_price=total_amount,
                line_total=total_amount,
            )
        return orders

    def import_technician_costs(self, workbook, users, orders, staff_by_name):
        if "تكلفة تشغيلية" not in workbook.sheetnames:
            return
        ws = workbook["تكلفة تشغيلية"]
        headers = [cell.value for cell in ws[1]]
        for row in ws.iter_rows(min_row=2, values_only=True):
            data = row_dict(headers, row)
            order = orders.get(clean_text(data.get("أمر ID")))
            technician = clean_text(data.get("الفني"))
            if not order or not technician:
                continue
            staff = staff_by_name.get(technician)
            if staff is None:
                staff, _ = StaffProfile.objects.get_or_create(
                    name=technician, defaults={"staff_role": "Excel technician"}
                )
                staff_by_name[technician] = staff
            task_name = clean_text(data.get("المهمة")) or "Excel task"
            cost = money(data.get("التكلفة"))
            JobStaffAssignment.objects.get_or_create(
                order=order,
                staff=staff,
                role=task_name,
                scheduled_start_at=as_aware(data.get("التاريخ")),
                scheduled_end_at=as_aware(data.get("التاريخ")) + timedelta(hours=8),
                defaults={"cost": cost, "status": "completed"},
            )
            StaffTask.objects.get_or_create(
                staff=staff,
                order=order,
                title=task_name,
                defaults={"status": "done", "due_at": as_aware(data.get("التاريخ")), "notes": "Excel import technician cost"},
            )

    def import_collections(self, workbook, users, orders):
        if "التحصيلات" not in workbook.sheetnames:
            return
        cashbox, _ = Cashbox.objects.get_or_create(
            name="Excel Collections Cashbox", defaults={"opening_balance": Decimal("0"), "currency": "EGP"}
        )
        headers = [cell.value for cell in workbook["التحصيلات"][1]]
        for row_number, row in enumerate(workbook["التحصيلات"].iter_rows(min_row=2, values_only=True), start=2):
            data = row_dict(headers, row)
            order_id = clean_text(data.get("أمر ID"))
            amount = money(data.get("القيمة"))
            if not amount:
                continue
            client, _ = Client.objects.get_or_create(name=clean_text(data.get("اسم العميل")) or "Excel Client")
            order = orders.get(order_id)
            invoice = getattr(order, "invoices", None).first() if order else None
            method_name = clean_text(data.get("طريقة التحصيل")) or "Cash"
            method, _ = PaymentMethod.objects.get_or_create(name=method_name, defaults={"method_type": "cash" if "cash" in method_name.lower() else "bank"})
            reference = f"EXCEL-COLLECTION-{row_number}"
            Payment.objects.update_or_create(
                reference=reference,
                defaults={
                    "client": client,
                    "order": order,
                    "invoice": invoice,
                    "amount": amount,
                    "payment_date": as_date(data.get("التاريخ")),
                    "method": method,
                    "cashbox": cashbox if method.method_type == "cash" else None,
                    "notes": f"Collector: {clean_text(data.get('المُحصل'))}",
                    "created_by": users[ROLE_ACCOUNTANT],
                },
            )
            if method.method_type == "cash":
                CashTransaction.objects.update_or_create(
                    cashbox=cashbox,
                    description=f"[Excel] Collection {reference}",
                    defaults={
                        "transaction_type": "in",
                        "amount": amount,
                        "transaction_date": as_date(data.get("التاريخ")),
                        "order": order,
                        "created_by": users[ROLE_CASHIER],
                    },
                )

    def import_vendor_driver_payments(self, workbook, users, vendors, drivers):
        if "الموردين" not in workbook.sheetnames:
            return
        ws = workbook["الموردين"]
        for row in ws.iter_rows(min_row=3, values_only=True):
            vendor_name = clean_text(row[0])
            if vendor_name:
                vendor, _ = Vendor.objects.get_or_create(name=vendor_name, defaults={"service_type": "Excel vendor"})
                debt = money(row[1])
                if debt:
                    VendorTransaction.objects.get_or_create(
                        vendor=vendor,
                        description=f"Excel supplier opening debt {vendor_name}",
                        defaults={
                            "transaction_type": "cost",
                            "amount": debt,
                            "transaction_date": timezone.localdate(),
                            "created_by": users[ROLE_ACCOUNTANT],
                        },
                    )
            pay_vendor_name = clean_text(row[7])
            pay_amount = money(row[6])
            if pay_vendor_name and pay_amount:
                vendor, _ = Vendor.objects.get_or_create(name=pay_vendor_name, defaults={"service_type": "Excel vendor"})
                VendorPayment.objects.get_or_create(
                    vendor=vendor,
                    amount=pay_amount,
                    payment_date=as_date(row[5]),
                    defaults={
                        "method": "cash",
                        "reference": f"EXCEL-VENDOR-{row[5]}-{pay_vendor_name}-{pay_amount}",
                        "created_by": users[ROLE_ACCOUNTANT],
                    },
                )
            driver_name = clean_text(row[12])
            driver_amount = money(row[11])
            if driver_name and driver_amount:
                driver, _ = Driver.objects.get_or_create(name=driver_name, defaults={"day_rate": Decimal("0")})
                DriverPayment.objects.get_or_create(
                    driver=driver,
                    amount=driver_amount,
                    payment_date=as_date(row[10]),
                    defaults={
                        "method": "cash",
                        "reference": f"EXCEL-DRIVER-{row[10]}-{driver_name}-{driver_amount}",
                        "created_by": users[ROLE_ACCOUNTANT],
                    },
                )

    def import_cash_custody(self, workbook, users, orders):
        if "عهدة" not in workbook.sheetnames:
            return
        cashbox, _ = Cashbox.objects.get_or_create(
            name="Excel Cash Custody", defaults={"opening_balance": Decimal("0"), "currency": "EGP"}
        )
        for row_number, row in enumerate(workbook["عهدة"].iter_rows(min_row=5, values_only=True), start=5):
            expense_date = as_date(row[0]) or timezone.localdate()
            category_name = clean_text(row[2])
            amount = money(row[4])
            if not category_name or not amount:
                continue
            category, _ = ExpenseCategory.objects.get_or_create(name=category_name)
            description = f"[Excel] Cash custody row {row_number}: {category_name}"
            Expense.objects.update_or_create(
                description=description,
                defaults={
                    "category": category,
                    "amount": amount,
                    "expense_date": expense_date,
                    "status": "approved",
                    "created_by": users[ROLE_ACCOUNTANT],
                },
            )
            CashTransaction.objects.update_or_create(
                cashbox=cashbox,
                description=description,
                defaults={
                    "transaction_type": "out",
                    "amount": amount,
                    "transaction_date": expense_date,
                    "created_by": users[ROLE_CASHIER],
                },
            )
